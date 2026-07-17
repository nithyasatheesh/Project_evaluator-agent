"""
exporter.py
===========
Builds the final Excel workbook of evaluation results: one row per
student, one column per rubric criterion (fully dynamic — never
hardcoded) showing the score achieved under that criterion's own
column, a live formula-driven Total and Percentage, grade, qualitative
feedback, separate Areas of Strength / Areas of Improvement columns,
and parser warnings.
"""

from __future__ import annotations

import io
from pathlib import Path

from utils import EvaluationResult, Rubric, RubricCriterion, sanitize_sheet_name, setup_logger

logger = setup_logger(__name__)

# Columns that hold longer free-text content and should get wrapped text
# and a wider column width in the exported sheet.
_WRAP_COLUMNS = {
    "Language Feedback",
    "Analysis Feedback",
    "Clarity Feedback",
    "Overall Feedback",
    "Areas of Strength",
    "Areas of Improvement",
    "Parser Warnings",
}

# Columns that are numeric scores/percentages and should get a fixed
# numeric display format rather than Excel's default "General" format.
_NUMERIC_FORMAT_COLUMNS = {"Total", "Max Score", "Percentage"}

_LIST_JOIN_SEPARATOR = "\n"


def _format_max(value: float) -> str:
    """Render a max score without a trailing '.0' for whole numbers."""
    return str(int(value)) if float(value).is_integer() else str(value)


def criterion_header(criterion: RubricCriterion) -> str:
    """Build the Excel column header for a rubric criterion.

    Includes the criterion's maximum score in the header itself (e.g.
    "Code Quality (Max 25)") so the score shown in each cell is always
    self-describing without needing a separate lookup column.
    """
    return f"{criterion.name} (Max {_format_max(criterion.max_score)})"


def _build_rows(results: list[EvaluationResult], rubric: Rubric) -> list[dict]:
    """Turn EvaluationResult objects into flat dict rows for a DataFrame."""
    rows: list[dict] = []
    for result in results:
        row: dict = {"Student": result.student}
        for criterion in rubric.criteria:
            row[criterion_header(criterion)] = result.scores.get(criterion.name, 0.0)
        # Total/Percentage are seeded with computed values here so the
        # DataFrame has a valid numeric column to write; both are then
        # replaced with live Excel formulas in _apply_formulas so the
        # workbook recalculates automatically if a score is edited.
        row["Total"] = round(result.total_score, 2)
        row["Max Score"] = round(result.max_score, 2)
        row["Percentage"] = round(result.percentage, 2)
        row["Grade"] = result.grade
        row["Language Feedback"] = result.language_feedback
        row["Analysis Feedback"] = result.analysis_feedback
        row["Clarity Feedback"] = result.clarity_feedback
        row["Overall Feedback"] = result.overall_feedback
        row["Areas of Strength"] = _LIST_JOIN_SEPARATOR.join(f"- {s}" for s in result.strengths)
        row["Areas of Improvement"] = _LIST_JOIN_SEPARATOR.join(f"- {s}" for s in result.improvements)
        row["Parser Warnings"] = _LIST_JOIN_SEPARATOR.join(result.warnings)
        row["Evaluation Failed"] = "Yes" if result.evaluation_failed else "No"
        rows.append(row)
    return rows


def _apply_formulas(sheet, rubric: Rubric) -> None:
    """Replace the static Total/Percentage values with live Excel formulas.

    Total becomes =SUM(<first criterion column>:<last criterion column>)
    for that row, and Percentage becomes that Total divided by Max Score
    (as a percentage), so both values recompute automatically in Excel
    if a faculty member manually edits any per-criterion score.
    """
    from openpyxl.utils import get_column_letter

    if not rubric.criteria:
        return

    headers = [cell.value for cell in sheet[1]]
    criterion_col_nums = [
        headers.index(criterion_header(c)) + 1 for c in rubric.criteria if criterion_header(c) in headers
    ]
    if not criterion_col_nums:
        return

    first_letter = get_column_letter(min(criterion_col_nums))
    last_letter = get_column_letter(max(criterion_col_nums))

    total_col = headers.index("Total") + 1 if "Total" in headers else None
    max_col = headers.index("Max Score") + 1 if "Max Score" in headers else None
    pct_col = headers.index("Percentage") + 1 if "Percentage" in headers else None
    total_letter = get_column_letter(total_col) if total_col else None
    max_letter = get_column_letter(max_col) if max_col else None

    for row_idx in range(2, sheet.max_row + 1):
        if total_col:
            sheet.cell(row=row_idx, column=total_col).value = (
                f"=SUM({first_letter}{row_idx}:{last_letter}{row_idx})"
            )
        if pct_col and total_letter and max_letter:
            sheet.cell(row=row_idx, column=pct_col).value = (
                f"=IFERROR(ROUND({total_letter}{row_idx}/{max_letter}{row_idx}*100,2),0)"
            )


def _style_workbook(buffer: io.BytesIO, sheet_name: str, rubric: Rubric) -> io.BytesIO:
    """Apply header styling, formulas, column widths and text wrapping in-place."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook[sheet_name]

    _apply_formulas(sheet, rubric)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    criterion_headers = {criterion_header(c) for c in rubric.criteria}

    headers = [cell.value for cell in sheet[1]]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

        column_letter = get_column_letter(col_idx)
        if header in _WRAP_COLUMNS:
            sheet.column_dimensions[column_letter].width = 45
        elif header == "Student":
            sheet.column_dimensions[column_letter].width = 22
        elif header in criterion_headers:
            sheet.column_dimensions[column_letter].width = 20
        else:
            sheet.column_dimensions[column_letter].width = 16

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            header = headers[cell.column - 1]
            if header in _WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
            if header in criterion_headers or header in _NUMERIC_FORMAT_COLUMNS:
                cell.number_format = "0.##"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    out_buffer = io.BytesIO()
    workbook.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer


def export_results_to_bytes(results: list[EvaluationResult], rubric: Rubric, sheet_name: str = "Evaluation Results") -> bytes:
    """Build the evaluation Excel workbook and return its raw bytes."""
    import pandas as pd

    if not results:
        raise ValueError("No evaluation results to export.")

    rows = _build_rows(results, rubric)
    df = pd.DataFrame(rows)

    safe_sheet_name = sanitize_sheet_name(sheet_name)
    raw_buffer = io.BytesIO()
    with pd.ExcelWriter(raw_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

    styled_buffer = _style_workbook(raw_buffer, safe_sheet_name, rubric)
    return styled_buffer.getvalue()


def export_results_to_file(results: list[EvaluationResult], rubric: Rubric, output_path: Path) -> Path:
    """Build the evaluation Excel workbook and write it to output_path."""
    data = export_results_to_bytes(results, rubric)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(data)
    logger.info("Wrote evaluation results workbook to '%s'", output_path)
    return output_path
